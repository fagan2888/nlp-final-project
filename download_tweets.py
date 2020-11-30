import os
import string
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

INPUT_PATH = os.path.join(os.path.realpath(os.path.dirname(__file__)), 'input.xlsx')
OUTPUT_PATH = os.path.join(os.path.realpath(os.path.dirname(__file__)), 'output.xlsx')

def check_load_xpath(driver, xp, timeout):
	try:
		e = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xp)))
		print('page ready')
		return 0
	except TimeoutError as err:
		print('connection timed out')
		print(err)
		time.sleep(5)
		return 1

def load_excel():
	search_requests = []
	wb = openpyxl.load_workbook(filename=INPUT_PATH)
	ws = wb.active
	for cell in ws['A1':'Z1']:
		for i in range(len(cell)):
			if(cell[i].value):
				search_requests.append(cell[i].value)
	return search_requests

def write_excel(col, row, text):
	if(col > 26):
		print('EXCEL WRITE INDEX OUT OF RANGE')
		return 1
	if(os.path.isfile(OUTPUT_PATH) and os.access(OUTPUT_PATH, os.R_OK)):
		print('output file exists and is readable')
		wb = openpyxl.load_workbook(filename=OUTPUT_PATH)
	else:
		print('output file not present, creating new file')
		wb = openpyxl.load_workbook(filename=INPUT_PATH)
	ws = wb.active
	ws[string.ascii_lowercase[col] + str(row)] = text
	wb.save('output.xlsx')


def main():
	req = load_excel()
	print(req)
	limit = 25
	driver = webdriver.Chrome(executable_path=r'C:\Users\james\Bin\chromedriver')
	driver.get('https://twitter.com/explore')
	for i in range(len(req)):

		if(check_load_xpath(driver, '/html/body/div/div/div/div[2]/header', 10)):
			return 1

		if(i > 0):
			driver.find_element_by_xpath('/html/body/div/div/div/div[2]/main/div/div/div/div/div/div[1]/div[1]/div/div/div/div/div[2]/div[2]/div/div/div/form/div[1]/div/div/div[2]/input').clear()
			driver.find_element_by_xpath('/html/body/div/div/div/div[2]/main/div/div/div/div/div/div[1]/div[1]/div/div/div/div/div[2]/div[2]/div/div/div/form/div[1]/div/div/div[2]/input').send_keys(req[i])
			driver.find_element_by_xpath('/html/body/div/div/div/div[2]/main/div/div/div/div/div/div[1]/div[1]/div/div/div/div/div[2]/div[2]/div/div/div/form/div[1]/div/div/div[2]/input').send_keys(Keys.RETURN)
		else:
			driver.find_element_by_xpath('/html/body/div/div/div/div[2]/main/div/div/div/div/div/div[1]/div[1]/div/div/div/div/div[1]/div[2]/div/div/div/form/div[1]/div/div/div[2]/input').send_keys(req[i])
			driver.find_element_by_xpath('/html/body/div/div/div/div[2]/main/div/div/div/div/div/div[1]/div[1]/div/div/div/div/div[1]/div[2]/div/div/div/form/div[1]/div/div/div[2]/input').send_keys(Keys.RETURN)
		time.sleep(1)

		if(check_load_xpath(driver, '/html/body/div/div/div/div[2]/main/div/div/div/div[1]/div/div[1]/div[2]/nav/div/div[2]/div/div[2]/a', 10)):
			return 1
		driver.find_element_by_xpath('/html/body/div/div/div/div[2]/main/div/div/div/div[1]/div/div[1]/div[2]/nav/div/div[2]/div/div[2]/a').click()
		if(check_load_xpath(driver, '/html/body/div/div/div/div[2]/main/div/div/div/div/div/div[2]/div/div/section/div/div/div[1]/div/div/article/div/div/div/div[2]', 10)):
			return 1
		for j in range(1,limit):
			if(j % 10 == 0):
				driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
				time.sleep(3)
			k = j - (15 * (j // 15))
			path = '/html/body/div/div/div/div[2]/main/div/div/div/div/div/div[2]/div/div/section/div/div/div[' + str(k) + ']/div/div/article/div/div/div/div[2]'
			try:
				if(driver.find_element_by_xpath(path).text):
					print(driver.find_element_by_xpath(path).text)
					write_excel(i, j+1, driver.find_element_by_xpath(path).text)
			except:
				continue

main()
